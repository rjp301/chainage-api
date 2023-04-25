from fastapi import UploadFile, File, HTTPException, Depends, APIRouter
from fastapi.responses import FileResponse

import shapefile
import openpyxl
import tempfile
import zipfile
import os

router = APIRouter(prefix="/convert")

def create_excel_workbook(file_path, worksheet_data):
  wb = openpyxl.Workbook()
  for sheet_name, sheet_data in worksheet_data.items():
    sheet = wb.create_sheet(sheet_name)
    for row_idx, row_data in enumerate(sheet_data):
      for col_idx, col_data in enumerate(row_data):
        sheet.cell(row=row_idx+1, column=col_idx+1, value=col_data)
  wb.save(file_path)
  return wb

async def get_temp_dir():
  fname = tempfile.NamedTemporaryFile(suffix=".xlsx")
  try: yield fname.name
  finally: del fname


@router.post("/json-to-excel")
async def json_to_excel(data: dict, fname=Depends(get_temp_dir)):
  filename = data["filename"] + ".xlsx"
  worksheet_data = data["worksheets"]
 
  create_excel_workbook(fname, worksheet_data)
  return FileResponse(fname, 
    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
    filename=filename
  )


@router.post("/shapefile-to-geojson")
async def shapefile_to_geojson(file: UploadFile = File(...)):
  with tempfile.TemporaryDirectory() as temp_dir:
    zip_path = os.path.join(temp_dir,file.filename)
    
    with open(zip_path,"wb") as f:
      f.write(await file.read())

    with zipfile.ZipFile(zip_path,"r") as zip_ref:
      zip_ref.extractall(temp_dir)   
    
    files = os.listdir(temp_dir)
    shp_file = next((f for f in files if f.endswith(".shp")),None)

    if not shp_file:
      raise HTTPException(status_code=400, detail="Uploaded zip file does not contain a .shp file")

    # Save uploaded file to temporary directory
    file_path = os.path.join(temp_dir,shp_file)

    # Check if the uploaded file is a valid shapefile
    try:
      sf = shapefile.Reader(str(file_path))
    except shapefile.ShapefileException:
      raise HTTPException(status_code=400, detail="Uploaded file is not a valid Shapefile")

    # Extract the shapefile's fields and features
    fields = [f[0] for f in sf.fields[1:]]
    features = []
    for sr in sf.shapeRecords():
      geom = sr.shape.__geo_interface__
      atr = dict(zip(fields, sr.record))
      features.append(dict(geometry=geom, properties=atr))

  return dict(type="FeatureCollection", features=features)
